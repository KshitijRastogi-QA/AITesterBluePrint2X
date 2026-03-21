import pandas as pd

# Creating a highly detailed, enterprise-grade test case repository mappings
# based strictly on the Restful Booker API documentation.

data = [
    # Auth - Create Token
    {
        "Scenario TID": "TC_AUTH_01",
        "TestCase Description": "Verify successful token generation with valid credentials.",
        "PreCondition": "API server is up and accessible over HTTPS. Valid credentials are known.",
        "Test Data": "Username: admin\nPassword: password123",
        "TestSteps": "1. Set URL to https://restful-booker.herokuapp.com/auth\n2. Set HTTP Method to POST\n3. Set Request Headers: Content-Type='application/json'\n4. Provide Request Body:\n{\n  \"username\": \"admin\",\n  \"password\": \"password123\"\n}\n5. Send Request",
        "Expected Result": "Expected Status Code: 200 OK\nExpected Response Time: < 2s\nExpected Response Headers: Content-Type='application/json'\nExpected Response Body: JSON object containing a valid alphanumeric 'token' string.",
        "Status": "Not Executed",
        "Executed QA Name ": "Kshitij Rastogi",
        "Priority": "P0",
        "Misc (Comments)": "Positive Scenario",
        "Is Automated": "Yes"
    },
    {
        "Scenario TID": "TC_AUTH_02",
        "TestCase Description": "Verify token generation failure with invalid password.",
        "PreCondition": "API server is up and accessible over HTTPS.",
        "Test Data": "Username: admin\nPassword: wrongpassword!",
        "TestSteps": "1. Set URL to https://restful-booker.herokuapp.com/auth\n2. Set HTTP Method to POST\n3. Set Request Headers: Content-Type='application/json'\n4. Provide Request Body:\n{\n  \"username\": \"admin\",\n  \"password\": \"wrongpassword!\"\n}\n5. Send Request",
        "Expected Result": "Expected Status Code: 200 OK (Per Restful-booker implementation, it returns 200 with an error object)\nExpected Response Time: < 2s\nExpected Response Body: {\"reason\": \"Bad credentials\"}",
        "Status": "Not Executed",
        "Executed QA Name ": "Kshitij Rastogi",
        "Priority": "P1",
        "Misc (Comments)": "Negative Scenario",
        "Is Automated": "Yes"
    },

    # Booking - GetBookingIds
    {
        "Scenario TID": "TC_BKG_GET_ALL_01",
        "TestCase Description": "Retrieve all booking IDs without any query parameters.",
        "PreCondition": "API server is up. At least one booking exists in the database.",
        "Test Data": "N/A",
        "TestSteps": "1. Set URL to https://restful-booker.herokuapp.com/booking\n2. Set HTTP Method to GET\n3. Send Request",
        "Expected Result": "Expected Status Code: 200 OK\nExpected Response Body: An array of JSON objects containing 'bookingid'.\nExample:\n[\n  {\"bookingid\": 1},\n  {\"bookingid\": 2}\n]",
        "Status": "Not Executed",
        "Executed QA Name ": "Kshitij Rastogi",
        "Priority": "P0",
        "Misc (Comments)": "Positive Scenario",
        "Is Automated": "Yes"
    },
    {
        "Scenario TID": "TC_BKG_GET_FILTER_02",
        "TestCase Description": "Filter booking IDs by firstname and lastname query parameters.",
        "PreCondition": "API server is up. A booking with firstname 'sally' and lastname 'brown' exists.",
        "Test Data": "Query Params: firstname=sally, lastname=brown",
        "TestSteps": "1. Set URL to https://restful-booker.herokuapp.com/booking?firstname=sally&lastname=brown\n2. Set HTTP Method to GET\n3. Send Request",
        "Expected Result": "Expected Status Code: 200 OK\nExpected Response Body: Array of JSON objects with 'bookingid' matching the query criteria.",
        "Status": "Not Executed",
        "Executed QA Name ": "Kshitij Rastogi",
        "Priority": "P1",
        "Misc (Comments)": "Positive Scenario",
        "Is Automated": "Yes"
    },

    # Booking - GetBooking
    {
        "Scenario TID": "TC_BKG_GET_SINGLE_01",
        "TestCase Description": "Retrieve a specific booking by its valid Booking ID.",
        "PreCondition": "Booking ID 1 exists in the system. API is accessible.",
        "Test Data": "Path Parameter id = 1",
        "TestSteps": "1. Set URL to https://restful-booker.herokuapp.com/booking/1\n2. Set HTTP Method to GET\n3. Set Request Headers: Accept='application/json'\n4. Send Request",
        "Expected Result": "Expected Status Code: 200 OK\nExpected Response Body: JSON object containing firstname, lastname, totalprice, depositpaid, bookingdates (checkin, checkout), and additionalneeds.",
        "Status": "Not Executed",
        "Executed QA Name ": "Kshitij Rastogi",
        "Priority": "P0",
        "Misc (Comments)": "Positive Scenario",
        "Is Automated": "Yes"
    },
    {
        "Scenario TID": "TC_BKG_GET_SINGLE_02",
        "TestCase Description": "Attempt to retrieve a booking that does not exist.",
        "PreCondition": "API is accessible.",
        "Test Data": "Path Parameter id = 9999999",
        "TestSteps": "1. Set URL to https://restful-booker.herokuapp.com/booking/9999999\n2. Set HTTP Method to GET\n3. Set Request Headers: Accept='application/json'\n4. Send Request",
        "Expected Result": "Expected Status Code: 404 Not Found\nExpected Response Time: < 2s\nExpected Response Body: 'Not Found' text or error standard payload.",
        "Status": "Not Executed",
        "Executed QA Name ": "Kshitij Rastogi",
        "Priority": "P1",
        "Misc (Comments)": "Negative Scenario",
        "Is Automated": "Yes"
    },

    # Booking - CreateBooking
    {
        "Scenario TID": "TC_BKG_CREATE_01",
        "TestCase Description": "Create a new booking with a fully valid JSON payload.",
        "PreCondition": "API is accessible.",
        "Test Data": "Firstname: Jim\nLastname: Brown\nTotalPrice: 111\nDepositPaid: true\nCheckin: 2018-01-01\nCheckout: 2019-01-01\nAdditionalNeeds: Breakfast",
        "TestSteps": "1. Set URL to https://restful-booker.herokuapp.com/booking\n2. Set HTTP Method to POST\n3. Set Request Headers: Content-Type='application/json', Accept='application/json'\n4. Provide Request Body:\n{\n    \"firstname\" : \"Jim\",\n    \"lastname\" : \"Brown\",\n    \"totalprice\" : 111,\n    \"depositpaid\" : true,\n    \"bookingdates\" : {\n        \"checkin\" : \"2018-01-01\",\n        \"checkout\" : \"2019-01-01\"\n    },\n    \"additionalneeds\" : \"Breakfast\"\n}\n5. Send Request",
        "Expected Result": "Expected Status Code: 200 OK\nExpected Response Body: Includes 'bookingid' integer and the nested 'booking' object verifying data consistency.",
        "Status": "Not Executed",
        "Executed QA Name ": "Kshitij Rastogi",
        "Priority": "P0",
        "Misc (Comments)": "Positive Scenario",
        "Is Automated": "Yes"
    },

    # Booking - UpdateBooking
    {
        "Scenario TID": "TC_BKG_UPDATE_01",
        "TestCase Description": "Perform a Full Update on an existing booking using valid Auth Cookie.",
        "PreCondition": "Booking ID 1 exists. An Auth token has been generated.",
        "Test Data": "Path Parameter id = 1\nCookie: token=abcd1234\nUpdate Data: James Brown, Price: 111",
        "TestSteps": "1. Set URL to https://restful-booker.herokuapp.com/booking/1\n2. Set HTTP Method to PUT\n3. Set Request Headers:\n  Content-Type='application/json'\n  Accept='application/json'\n  Cookie='token=abcd1234'\n4. Provide Request Body:\n{\n    \"firstname\" : \"James\",\n    \"lastname\" : \"Brown\",\n    \"totalprice\" : 111,\n    \"depositpaid\" : true,\n    \"bookingdates\" : {\n        \"checkin\" : \"2018-01-01\",\n        \"checkout\" : \"2019-01-01\"\n    },\n    \"additionalneeds\" : \"Breakfast\"\n}\n5. Send Request",
        "Expected Result": "Expected Status Code: 200 OK\nExpected Response Body: Returns the fully updated 'booking' JSON object mapping to the provided request payload.",
        "Status": "Not Executed",
        "Executed QA Name ": "Kshitij Rastogi",
        "Priority": "P0",
        "Misc (Comments)": "Positive Scenario",
        "Is Automated": "Yes"
    },
    {
        "Scenario TID": "TC_BKG_UPDATE_02",
        "TestCase Description": "Perform a Full Update on an existing booking without any Auth token.",
        "PreCondition": "Booking ID 1 exists.",
        "Test Data": "Path Parameter id = 1\nNo Cookie/Authorization Headers applied",
        "TestSteps": "1. Set URL to https://restful-booker.herokuapp.com/booking/1\n2. Set HTTP Method to PUT\n3. Provide Request Body with valid JSON data\n4. Send Request",
        "Expected Result": "Expected Status Code: 403 Forbidden\nExpected Response Body: Indicates Authorization failure.",
        "Status": "Not Executed",
        "Executed QA Name ": "Kshitij Rastogi",
        "Priority": "P1",
        "Misc (Comments)": "Negative Scenario",
        "Is Automated": "Yes"
    },

    # Booking - PartialUpdateBooking
    {
        "Scenario TID": "TC_BKG_PARTIAL_01",
        "TestCase Description": "Perform a Partial Update (PATCH) on firstname and lastname with valid Basic Auth.",
        "PreCondition": "Booking ID 1 exists.",
        "Test Data": "Path Parameter id = 1\nAuthorization Header = Basic YWRtaW46cGFzc3dvcmQxMjM=",
        "TestSteps": "1. Set URL to https://restful-booker.herokuapp.com/booking/1\n2. Set HTTP Method to PATCH\n3. Set Request Headers:\n  Content-Type='application/json'\n  Accept='application/json'\n  Authorization='Basic YWRtaW46cGFzc3dvcmQxMjM='\n4. Provide Request Body:\n{\n    \"firstname\" : \"James\",\n    \"lastname\" : \"Brown\"\n}\n5. Send Request",
        "Expected Result": "Expected Status Code: 200 OK\nExpected Response Body: Returns the full updated booking object where only firstname ('James') and lastname ('Brown') are altered.",
        "Status": "Not Executed",
        "Executed QA Name ": "Kshitij Rastogi",
        "Priority": "P0",
        "Misc (Comments)": "Positive Scenario",
        "Is Automated": "Yes"
    },

    # Booking - DeleteBooking
    {
        "Scenario TID": "TC_BKG_DELETE_01",
        "TestCase Description": "Delete an existing booking with a valid authentication cookie.",
        "PreCondition": "Booking ID 1 exists. Auth token is valid.",
        "Test Data": "Path Parameter id = 1\nCookie: token=xyz987",
        "TestSteps": "1. Set URL to https://restful-booker.herokuapp.com/booking/1\n2. Set HTTP Method to DELETE\n3. Set Request Headers:\n  Cookie='token=xyz987'\n4. Send Request",
        "Expected Result": "Expected Status Code: 201 Created (Per Restful-booker standard for DELETE success)\nExpected Response Body: Blank or standard success indicator.",
        "Status": "Not Executed",
        "Executed QA Name ": "Kshitij Rastogi",
        "Priority": "P0",
        "Misc (Comments)": "Positive Scenario",
        "Is Automated": "Yes"
    },

    # Ping - HealthCheck
    {
        "Scenario TID": "TC_PING_01",
        "TestCase Description": "Verify the HealthCheck Ping endpoint validates server uptime.",
        "PreCondition": "None",
        "Test Data": "N/A",
        "TestSteps": "1. Set URL to https://restful-booker.herokuapp.com/ping\n2. Set HTTP Method to GET\n3. Send Request",
        "Expected Result": "Expected Status Code: 201 Created\nExpected Response Time: < 2s\nExpected Response Body: Blank or 'Created' string.",
        "Status": "Not Executed",
        "Executed QA Name ": "Kshitij Rastogi",
        "Priority": "P2",
        "Misc (Comments)": "Positive Scenario",
        "Is Automated": "Yes"
    }
]

df_cases = pd.DataFrame(data)

# Exact Template columns constraint requested by user
template_cols = [
    'Scenario TID', 'Test Data', 'TestCase Description', 'PreCondition', 
    'TestSteps', 'Expected Result', 'Actual Result', 'Steps to Execute', 
    'Expected Result.1', 'Actual Result.1', 'Status', 'Executed QA Name ', 
    'Misc (Comments)', 'Priority', 'Is Automated'
]

# Map over data to final dataframe
df_final = pd.DataFrame(columns=template_cols)
for col in template_cols:
    if col in df_cases.columns:
        df_final[col] = df_cases[col]
    elif col == 'Steps to Execute':
        df_final[col] = df_cases['TestSteps']  # Handling duplicated semantic mapping

# Stylizing the Excal file for "Enterprise Grade" Appearance using an engine
filename = 'Restful_Booker_Enterprise_Test_Cases.xlsx'
with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    df_final.to_excel(writer, index=False, sheet_name='Restful Booker APIs')
    
    # Get the workbook and the specific worksheet
    workbook = writer.book
    worksheet = writer.sheets['Restful Booker APIs']
    
    # Freeze the top row
    worksheet.freeze_panes = "A2"
    
    # Formatting
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    regular_alignment = Alignment(wrap_text=True, vertical='top')

    # Apply headers style
    for cell in worksheet["1:1"]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Apply data alignment and column dimensions
    for col_idx in range(1, worksheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        
        # Expand widths intelligently
        if df_final.columns[col_idx-1] in ['TestSteps', 'Expected Result']:
            worksheet.column_dimensions[col_letter].width = 65
        elif df_final.columns[col_idx-1] in ['TestCase Description', 'PreCondition', 'Test Data']:
            worksheet.column_dimensions[col_letter].width = 40
        else:
            worksheet.column_dimensions[col_letter].width = 15

        for row_idx in range(2, worksheet.max_row + 1):
            worksheet[col_letter + str(row_idx)].alignment = regular_alignment

print(f"Beautifully formatted Excel test cases smoothly generated down path: {filename}")
